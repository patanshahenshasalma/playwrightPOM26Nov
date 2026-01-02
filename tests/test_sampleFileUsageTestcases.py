import json
import csv
import os
from dotenv import load_dotenv
from utils.readingJson import readJsonData
from utils.handlingCsv import insert_data_to_csv, read_dataFromCSV
from openpyxl import load_workbook
# from pydantic import BaseModel, ValidationError, validator
import pytest
from utils.pydanticUsage import ValidationError, User
from utils.handlingExcel import read_excel

searchData = "testData/searchData.json"
# testData = readJsonData(searchData)
testdata ="hello"

# @pytest.mark.parametrize("testData,tesdat2", [("hello", "hi"), ("hello2", "h12") ])
# def test_searchDataValues(testData, tesdat2):
#     print("testData",testData)
#     print("testData2", tesdat2)


# def test_dataFromCSV():
#     data = []
#     with open("testData\\credentails.csv") as csvFile:
#          datainCSV = csv.DictReader(csvFile)
#          for row in datainCSV:
#               data.append(row)
#     print(data)

# def test_writing():
#     with open("testData\\credentails.csv", mode="w", newline="") as csvFile:
#         writer = csv.DictWriter(csvFile, fieldnames=["username","password","age"])
#         writer.writeheader()
#         writer.writerow({ 
#             "username": "test1@gmail.com",
#             "password": "Admin1@123",
#             "age": 20
#             })

        

# def insert_data_to_csv():
#     with open("testData\\credentails.csv", mode="w", newline="") as csvFile:
#         fieldnames = ["username", "password", "age"]
#         writer = csv.DictWriter(csvFile, fieldnames=fieldnames)

#         writer.writeheader()

#         writer.writerow({
#             "username": "test@gmail.com",
#             "password": "Admin@123",
#             "age": 20
#         })

# def test_csvInsertion():
#     insert_data_to_csv()

# def test_printExcel():
#     from utils.handlingExcel import read_excel
#     filePath = "testData\\searchProduct.xlsx"
#     sheetName = "phoneConfigs"
#     data = read_excel(filePath, sheetName)
#     print(data)

# def test_printExcelvalues():
#     # workbook = load_workbook("testData\\searchProduct.xlsx")
#     # sheet = workbook["mixerDetails"]
#     # sheet.cell(column=1, row=3, value="netripro2")
#     # workbook.save("testData\\searchProduct.xlsx")
#     print(read_excel("testData\\searchProduct.xlsx","mixerDetails"))


# def test_writeExcel():
#     from utils.handlingExcel import addData_excel
#     filePath = "testData\\searchProduct.xlsx"
#     sheetName = "phoneConfigs"
#     value = "OnePlus"
#     addData_excel(filePath, sheetName, value)

# def test_append_to_excel():
#     from utils.handlingExcel import append_to_excel
#     file_path = "testData\\searchProduct.xlsx"
#     sheetName = "phoneConfigs"
#     headers = ["username", "password", "age"]
#     rows = [
#         ["user3@gmail.com", "Welcome@03", 35]
#     ]
#     append_to_excel(file_path, sheetName, headers, rows)

# def test_getDataFromCommandline():
#     env = os.getenv("env", "test")
#     # set env=test && pytest -s
#     load_dotenv(f".env.{env}") #.env.test
#     userName = os.getenv("Amazon_userName")
#     password =os.getenv("Amazon_pw")
#     print(userName)
#     print(password)


# def test_pydanticValidation_normalInput():
#     data = {
#         "username": "test@gmail.com",
#         "password": "Welcome@02",
#         "age": ""
#     }
#     user = User(**data)
#     print(user)

# def test_pydanticValidation_Json():
#      data = readJsonData("testData\\credentials.json")
#      user = User(**data)

# def test_pydanticValidation_csv():
#     rows = read_dataFromCSV("testData\\credentails.csv")
#     users = []

#     for i, row in enumerate(rows, start=1):
#         user = User(**row)   # fails automatically if invalid
#         users.append(user)