# pip install openpyxl
from openpyxl import load_workbook

def read_excel(filePath, sheetName):
    workbook = load_workbook(filePath)
    sheet = workbook[sheetName]
    data = []
    for row in sheet.iter_cols(min_col=2, values_only=True):
        data.append(row)
    return data




def addData_excel(filePath, sheetName, value):
    workbook = load_workbook(filePath)
    sheet = workbook[sheetName]
    sheet.cell(row=4, column=1, value=value)
    workbook.save(filePath)



def append_to_excel(file_path, sheetName, headers, rows):
    workbook = load_workbook(file_path)
    sheet = workbook[sheetName]
    sheet.append(headers)
    for row in rows:
        sheet.append(row)
    workbook.save(file_path)

# Usage example:
# append_to_excel(
#     "testData/users.xlsx","sheet",
#     ["user3@gmail.com", "Welcome@03", 35]
# )
